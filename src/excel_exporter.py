from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.aggregator import GEO_REGIONS, build_monthly_table, calculate_grand_totals, calculate_regional_totals
from src.logger import get_logger
from src.models import ClassifiedPayment

log = get_logger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="2D4A7A")
_TOTAL_FILL = PatternFill("solid", fgColor="E8EEF7")
_GRAND_FILL = PatternFill("solid", fgColor="C5D3EA")
_WHITE = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FONT = Font(bold=True, size=10)
_BODY_FONT = Font(size=10)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_header(ws, row_num: int, values: list[str]):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER


def _apply_row(ws, row_num: int, values: list, is_total: bool = False, is_grand: bool = False):
    fill = _GRAND_FILL if is_grand else (_TOTAL_FILL if is_total else _WHITE)
    font = _TOTAL_FONT if (is_total or is_grand) else _BODY_FONT
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.border = _BORDER
        if isinstance(val, float):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center")


def _write_geo_sheet(
    ws,
    payments: list[ClassifiedPayment],
    geo_region: str,
    year: Optional[int] = None,
    quarter: Optional[int] = None,
):
    headers = [
        "Month", "Month Start", "Month End",
        "Coaching", "Newsletter", "Illustrations", "Total Income",
        "Coaching Fee", "Newsletter Fee", "Illustrations Fee", "Total Fee",
    ]
    _apply_header(ws, 1, headers)

    rows = build_monthly_table(payments, geo_region, year, quarter)  # type: ignore[arg-type]
    for i, row in enumerate(rows, 2):
        is_total = row["Month"] == "TOTAL"
        vals = [
            row["Month"], row["Month Start"], row["Month End"],
            row["Coaching"], row["Newsletter"], row["Illustrations"], row["Total Income"],
            row["Coaching Fee"], row["Newsletter Fee"], row["Illustrations Fee"], row["Total Fee"],
        ]
        _apply_row(ws, i, vals, is_total=is_total)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.freeze_panes = "A2"


def _write_calculations_sheet(ws, payments: list[ClassifiedPayment]):
    headers = [
        "Region",
        "Coaching", "Newsletter", "Illustrations", "Total Income",
        "Coaching Fee", "Newsletter Fee", "Illustrations Fee", "Total Fee",
    ]
    _apply_header(ws, 1, headers)

    regional = calculate_regional_totals(payments)
    grand = calculate_grand_totals(payments)

    region_labels = {
        "SPAIN": "Spain",
        "EU_NOT_SPAIN": "EU (not Spain)",
        "OUTSIDE_EU": "Outside EU",
    }

    row_num = 2
    for region_key in GEO_REGIONS:
        d = regional.get(region_key, {})
        vals = [
            region_labels[region_key],
            round(d.get("coaching", 0), 2),
            round(d.get("newsletter", 0), 2),
            round(d.get("illustrations", 0), 2),
            round(d.get("total_income", 0), 2),
            round(d.get("coaching_fee", 0), 2),
            round(d.get("newsletter_fee", 0), 2),
            round(d.get("illustrations_fee", 0), 2),
            round(d.get("total_fee", 0), 2),
        ]
        _apply_row(ws, row_num, vals)
        row_num += 1

    grand_vals = [
        "GRAND TOTAL",
        round(grand.get("coaching", 0), 2),
        round(grand.get("newsletter", 0), 2),
        round(grand.get("illustrations", 0), 2),
        round(grand.get("total_income", 0), 2),
        round(grand.get("coaching_fee", 0), 2),
        round(grand.get("newsletter_fee", 0), 2),
        round(grand.get("illustrations_fee", 0), 2),
        round(grand.get("total_fee", 0), 2),
    ]
    _apply_row(ws, row_num, grand_vals, is_grand=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_import_sheet(ws, payments: list[ClassifiedPayment]):
    headers = [
        "id", "Created Date", "Description", "Converted Amount", "Converted Amount Refunded",
        "Fee", "Currency", "Net Amount",
        "IND_COACHING", "IND_NEWSLETTER", "IND_ILLUSTRATIONS",
        "IND_SPAIN", "IND_OUT_SPAIN", "IND_EXEU", "IND_EU",
        "Activity Type", "Geo Region", "Quarter", "Month", "Year",
        "Classification Rule", "Geo Rule",
    ]
    _apply_header(ws, 1, headers)

    for i, p in enumerate(payments, 2):
        vals = [
            p.id,
            p.created_date.strftime("%Y-%m-%d %H:%M:%S"),
            p.description,
            p.converted_amount,
            p.converted_amount_refunded,
            p.fee,
            p.currency.upper(),
            p.net_amount,
            p.IND_COACHING, p.IND_NEWSLETTER, p.IND_ILLUSTRATIONS,
            p.IND_SPAIN, p.IND_OUT_SPAIN, p.IND_EXEU, p.IND_EU,
            p.activity_type,
            p.geo_region,
            p.quarter,
            p.month,
            p.year,
            p.classification_rule,
            p.geo_rule,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    widths = [20, 20, 40, 15, 20, 10, 10, 12] + [12] * 7 + [15, 15, 10, 8, 8, 25, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def create_excel_report(
    payments: list[ClassifiedPayment],
    output_path: str | Path,
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    label: str = "",
) -> Path:
    """Generate the full Excel workbook and save to output_path."""
    wb = Workbook()

    ws_calc = wb.active
    ws_calc.title = "calculations"
    _write_calculations_sheet(ws_calc, payments)

    for region_key, sheet_name in [
        ("SPAIN", "calculations_Spain"),
        ("EU_NOT_SPAIN", "calculations_EUnotSpain"),
        ("OUTSIDE_EU", "calculations_EXEU"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_geo_sheet(ws, payments, region_key, year, quarter)

    ws_import = wb.create_sheet("import")
    _write_import_sheet(ws_import, payments)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("ℹ️ Excel report saved: %s", output_path)
    return output_path


def generate_report_filename(year: int, quarter: Optional[int] = None) -> str:
    if quarter:
        return f"Stripe_Report_Q{quarter}_{year}.xlsx"
    return f"Stripe_Report_{year}_Full.xlsx"
